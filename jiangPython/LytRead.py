import xlrd
import xlwt
from openpyxl import load_workbook
from xlutils.copy import copy

def read_excel(readUrl, writeUrl, date):

    #读你要算的excel
    workbook = xlrd.open_workbook(readUrl)
    sheet1 = workbook.sheet_by_index(0)

    #读你要写的excel
    workbook1 = xlrd.open_workbook(writeUrl)
    sheet2 = workbook1.sheet_by_index(0)

    Ecr = 0
    Hwr = 0
    countEcr = 0
    countHwr = 0
    for x in range(sheet1.nrows):
        if x == sheet1.nrows - 1:
            break
        data = sheet1.cell(x+1, 1).value
        if 'ECR' in data :
            Ecr = Ecr + sheet1.cell(x+1, 2).value
            countEcr = countEcr + 1
        if 'HWR' in data :
            Hwr = Hwr + sheet1.cell(x+1, 2).value
            countHwr = countHwr +1

    write = load_workbook(writeUrl)
    sheet3 = write.active
    sheet3.cell(sheet2.nrows+1, 1, date)
    sheet3.cell(sheet2.nrows+1, 2, Hwr)
    sheet3.cell(sheet2.nrows+1, 4, countHwr)
    sheet3.cell(sheet2.nrows+1, 8, Ecr)
    sheet3.cell(sheet2.nrows+1, 10, countEcr)
    write.save(writeUrl)




    # write = copy(workbook1)
    # sheet3 = write.get_sheet(0)
    # style = xlwt.easyxf('font:height 180, color-index black;align: wrap on, vert centre, horiz center')
    #
    # sheet3.write(sheet2.nrows+1, 1, date, style)
    # sheet3.write(sheet2.nrows+1, 2, Hwr, style)
    # sheet3.write(sheet2.nrows+1, 4, countHwr, style)
    # sheet3.write(sheet2.nrows+1, 8, Ecr, style)
    # sheet3.write(sheet2.nrows+1, 10, countEcr, style)
    # write.save(r'/Users/katyjiang/Desktop/ae8f7afa598fe220.xls')





if __name__ == '__main__':
    readUrl = r'/Users/katyjiang/Desktop/6.19mr.xls'
    writeUrl = r'/Users/katyjiang/Desktop/ae8f7afa598fe220.xlsx'
    date = "2018/6/19"
    read_excel(readUrl, writeUrl, date)


